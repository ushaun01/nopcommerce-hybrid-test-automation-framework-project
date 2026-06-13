import openpyxl
from openpyxl.styles import PatternFill



def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,row_num,column_no):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(row_num,column_no).value

def writeData(file,sheetName,row_num,column_no,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row_num,column_no).value = data
    workbook.save(file)

def fill_green(file,sheetName,row_num,column_no):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    green=PatternFill(start_color='00FF00',end_color='00FF00',patternType='solid')
    sheet.cell(row_num,column_no).fill=green
    workbook.save(file)

def fill_red(file,sheetName,row_num,column_no):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    red=PatternFill(start_color='FF0000',end_color='FF0000',patternType='solid')
    sheet.cell(row_num,column_no).fill=red
    workbook.save(file)